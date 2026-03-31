from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


OUTPUT_FILE = Path(__file__).resolve().parents[1] / "monthly cost.xlsx"


def usd(cell) -> None:
    cell.number_format = '"$"#,##0.00'


def write_headers(ws, row: int, headers: list[str]) -> None:
    fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    for idx, value in enumerate(headers, start=1):
        c = ws.cell(row=row, column=idx, value=value)
        c.font = Font(bold=True)
        c.fill = fill


def build_workbook() -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Monthly Estimate"

    today = date.today()
    first_day_month = date(today.year, today.month, 1)

    ws["A1"] = "CBS Azure Monthly Cost Estimate"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A3"] = "Pricing month"
    ws["B3"] = first_day_month
    ws["B3"].number_format = "yyyy-mm-dd"
    ws["A4"] = "Generated on"
    ws["B4"] = today
    ws["B4"].number_format = "yyyy-mm-dd"
    ws["A5"] = "Region"
    ws["B5"] = "West Europe"
    ws["A6"] = "Billing model"
    ws["B6"] = "Pay-as-you-go (Linux)"
    ws["A7"] = "Hours per month"
    ws["B7"] = 730
    ws["A8"] = "Currency"
    ws["B8"] = "USD"

    ws["A10"] = "Unit price assumptions"
    ws["A10"].font = Font(bold=True)
    write_headers(ws, 11, ["Resource", "Unit Price", "Unit"])

    ws["A12"] = "VM Standard_D4s_v5"
    ws["B12"] = 0.19
    usd(ws["B12"])
    ws["C12"] = "USD/hour"

    ws["A13"] = "VM Standard_D8s_v5"
    ws["B13"] = 0.38
    usd(ws["B13"])
    ws["C13"] = "USD/hour"

    ws["A14"] = "Premium SSD P20"
    ws["B14"] = 19
    usd(ws["B14"])
    ws["C14"] = "USD/month"

    ws["A15"] = "Standard Public IP"
    ws["B15"] = 3
    usd(ws["B15"])
    ws["C15"] = "USD/month"

    ws["A16"] = "Standard Load Balancer"
    ws["B16"] = 20
    usd(ws["B16"])
    ws["C16"] = "USD/month"

    ws["A17"] = "Application Gateway WAF_v2"
    ws["B17"] = 330
    usd(ws["B17"])
    ws["C17"] = "USD/month"

    ws["A19"] = "VM + Load Balancer architecture"
    ws["A19"].font = Font(bold=True)
    write_headers(
        ws,
        20,
        [
            "Environment",
            "D4 VM Count",
            "D8 VM Count",
            "P20 Disk Count",
            "Public IP Count",
            "LB Count",
            "Compute Monthly",
            "Disk Monthly",
            "Network Monthly",
            "Total Monthly",
        ],
    )

    def write_arch_row(row: int, name: str, d4: int, d8: int, disks: int, pips: int, lbs: int) -> None:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=d4)
        ws.cell(row=row, column=3, value=d8)
        ws.cell(row=row, column=4, value=disks)
        ws.cell(row=row, column=5, value=pips)
        ws.cell(row=row, column=6, value=lbs)
        ws.cell(row=row, column=7, value=f"=(B{row}*$B$12+C{row}*$B$13)*$B$7")
        ws.cell(row=row, column=8, value=f"=D{row}*$B$14")
        ws.cell(row=row, column=9, value=f"=E{row}*$B$15+F{row}*$B$16")
        ws.cell(row=row, column=10, value=f"=G{row}+H{row}+I{row}")
        for col in (7, 8, 9, 10):
            usd(ws.cell(row=row, column=col))

    write_arch_row(21, "Dev / Demo", 1, 0, 1, 1, 0)
    write_arch_row(22, "Staging", 2, 0, 2, 1, 1)
    write_arch_row(23, "Production small", 0, 3, 3, 1, 1)

    ws["A25"] = "VMSS + Application Gateway architecture"
    ws["A25"].font = Font(bold=True)
    write_headers(
        ws,
        26,
        [
            "Environment",
            "D4 VM Count",
            "D8 VM Count",
            "P20 Disk Count",
            "App Gateway Count",
            "Public IP Count",
            "Compute Monthly",
            "Disk Monthly",
            "Gateway+IP Monthly",
            "Total Monthly",
        ],
    )

    def write_vmss_row(row: int, name: str, d4: int, d8: int, disks: int, appgws: int, pips: int) -> None:
        ws.cell(row=row, column=1, value=name)
        ws.cell(row=row, column=2, value=d4)
        ws.cell(row=row, column=3, value=d8)
        ws.cell(row=row, column=4, value=disks)
        ws.cell(row=row, column=5, value=appgws)
        ws.cell(row=row, column=6, value=pips)
        ws.cell(row=row, column=7, value=f"=(B{row}*$B$12+C{row}*$B$13)*$B$7")
        ws.cell(row=row, column=8, value=f"=D{row}*$B$14")
        ws.cell(row=row, column=9, value=f"=E{row}*$B$17+F{row}*$B$15")
        ws.cell(row=row, column=10, value=f"=G{row}+H{row}+I{row}")
        for col in (7, 8, 9, 10):
            usd(ws.cell(row=row, column=col))

    write_vmss_row(27, "Staging", 2, 0, 2, 1, 1)
    write_vmss_row(28, "Production small", 0, 3, 3, 1, 1)
    write_vmss_row(29, "Production medium", 0, 6, 6, 1, 1)

    ws["A31"] = "Note"
    ws["B31"] = (
        "This workbook is generated monthly on day 1 by GitHub Actions. "
        "Review Azure Pricing Calculator for final purchase values."
    )

    ws.column_dimensions["A"].width = 34
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 17
    ws.column_dimensions["H"].width = 13
    ws.column_dimensions["I"].width = 19
    ws.column_dimensions["J"].width = 16

    return wb


def main() -> None:
    wb = build_workbook()
    wb.save(OUTPUT_FILE)
    print(f"Created {OUTPUT_FILE}")


if __name__ == "__main__":
    main()